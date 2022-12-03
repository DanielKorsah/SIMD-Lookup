import sqlite3
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from tqdm import tqdm, trange
import time




def populate_zone_ranks_table():
    sheet = postcode_workbook["All postcodes"]
    c = conn.cursor()

    percentage = 0
    total = 157834
    for row in tqdm(range(1700, total), ncols=100, colour="green", bar_format='{desc}: {percentage:3.2f}%|{bar}{r_bar}'):
        row_vals = []
        for col in range(1, 7):
            col_letter = get_column_letter(col)
            cell = f"{col_letter}{row}"
            row_vals.append(sheet[cell].value)
        c.execute("""
                        --sql
                        REPLACE INTO zone_ranks(postcode, data_zone, absolute_rank, rank_vigintile, rank_decile, rank_quintile)
                        VALUES (:postcode, :zone, :rank, :rank_vigintile, :rank_decile, :rank_quintile)
                        --endsql
                    """, {"postcode": row_vals[0], "zone": row_vals[1], "rank": row_vals[2], "rank_vigintile": row_vals[3], "rank_decile": row_vals[4], "rank_quintile": row_vals[5]})
        conn.commit()

def populate_zone_data_table():
    sheet = data_workbook["SIMD 2020v2 DZ lookup data"]
    c = conn.cursor()

    percentage = 0
    total = 6978
    for row in tqdm(range(2, total), ncols=100, colour="green", bar_format='{desc}: {percentage:3.2f}%|{bar}{r_bar}'):
        row_vals = []
        for col in [1, 2, 7, 8, 9, 10, 11, 12, 13, 14, 15]:
            col_letter = get_column_letter(col)
            cell = f"{col_letter}{row}"
            row_vals.append(sheet[cell].value)
        c.execute("""
                        --sql
                        REPLACE INTO zone_data(data_zone_id, data_zone_name, income, employment, education, health, access, crime, housing, total_population, working_age_population)
                        VALUES (:data_zone_id, :data_zone_name, :income, :employment, :education, :health, :access, :crime, :housing, :total_population, :working_age_population)
                        --endsql
                    """, {
                            "data_zone_id": row_vals[0],
                            "data_zone_name": row_vals[1],
                            "income": row_vals[2],
                            "employment": row_vals[3],
                            "education": row_vals[4],
                            "health": row_vals[5],
                            "access": row_vals[6],
                            "crime": row_vals[7],
                            "housing": row_vals[8],
                            "total_population": row_vals[9],
                            "working_age_population": row_vals[10],
                          }
                  )
        conn.commit()



conn = sqlite3.connect("simd_data.sqlite")
postcode_workbook = load_workbook("SIMD_sources/2020v2_postcode_lookup.xlsx")
print("load wb 1")
print(postcode_workbook.sheetnames)
populate_zone_ranks_table()

data_workbook = load_workbook("SIMD_sources/2020v2_datazone_lookup.xlsx")
print("load wb 2")
print(data_workbook.sheetnames)
populate_zone_data_table()






